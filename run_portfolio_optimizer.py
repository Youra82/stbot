#!/usr/bin/env python3
"""
run_portfolio_optimizer.py  (stbot)

Lädt alle Configs, führt Portfolio-Simulation (gemeinsamer Kapital-Pool,
kombinierte Equity-Kurve, echter MaxDD) durch und wählt das beste Portfolio
per Greedy-Algorithmus. Schreibt active_strategies in settings.json.

Aufruf:
  python3 run_portfolio_optimizer.py              # interaktiv
  python3 run_portfolio_optimizer.py --auto-write # automatisch (Scheduler)
"""
import contextlib
import io
import os
import sys
import json
import logging
import argparse
from datetime import date, timedelta
from tqdm import tqdm

# Ohne konfigurierten Handler verschluckt Python logger.info() komplett --
# siehe optimizer.py fuer Details (gleiches Problem, gleicher Fix).
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s', datefmt='%H:%M:%S')

PROJECT_ROOT  = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))

CONFIGS_DIR   = os.path.join(PROJECT_ROOT, 'src', 'stbot', 'strategy', 'configs')
SETTINGS_PATH = os.path.join(PROJECT_ROOT, 'settings.json')

B  = '\033[1;37m'
G  = '\033[0;32m'
Y  = '\033[1;33m'
R  = '\033[0;31m'
NC = '\033[0m'

DEFAULT_LOOKBACK_DAYS = 1095  # ~3 Jahre als Standard


def _scan_configs() -> list:
    if not os.path.isdir(CONFIGS_DIR):
        return []
    return sorted([
        os.path.join(CONFIGS_DIR, f)
        for f in os.listdir(CONFIGS_DIR)
        if f.endswith('.json')
    ])


def _build_strategies_data(config_files: list, start_date: str, end_date: str) -> dict:
    from stbot.analysis.backtester import load_data, FINE_TF_MAP, LazyFineData
    strategies_data = {}
    for path in tqdm(config_files, desc='Lade Configs & Daten'):
        fname = os.path.basename(path)
        try:
            with open(path) as f:
                config = json.load(f)
            market    = config.get('market', {})
            symbol    = market.get('symbol', '')
            timeframe = market.get('timeframe', '')
            htf       = market.get('htf')
            if not symbol or not timeframe:
                continue

            # Configs, die der Optimizer explizit als NICHT durch die IS/OOS-
            # Bestaetigung gelaufen markiert hat (_meta.confirmed == False),
            # fliessen nicht in die automatische Portfolio-Auswahl ein -- sonst
            # kann eine Kombination, die bereits als "haette live Geld verloren"
            # erkannt wurde, trotzdem in settings.json landen (live beobachtet:
            # 2026-08-21, unbestaetigtes BTC/6h mit OOS -20.1% wurde ins Portfolio
            # gewaehlt, weil dieser Optimizer eine eigene, unabhaengige Bewertung
            # ohne IS/OOS-Split macht und den confirmed-Status ignorierte).
            # Configs OHNE das Feld (aeltere, vor der IS/OOS-Aenderung erzeugte)
            # bleiben zugelassen, damit bestehende funktionierende Strategien
            # nicht ploetzlich rausfallen.
            meta = config.get('_meta', {})
            if meta.get('confirmed') is False:
                print(f"  {Y}Uebersprungen (IS/OOS nicht bestaetigt): {fname}{NC}")
                continue

            data = load_data(symbol, timeframe, start_date, end_date)
            if data is None or data.empty or len(data) < 50:
                print(f"  {Y}Uebersprungen (keine Daten): {fname}{NC}")
                continue

            # Feinere Kerzen fuer Intrabar-Pfad-Aufloesung (oraclebot-Muster) --
            # on-demand (LazyFineData): laedt nur die Tage, an denen im
            # Portfolio-Backtest tatsaechlich eine offene Position liegt,
            # statt den ganzen Zeitraum vorab herunterzuladen.
            fine_tf = FINE_TF_MAP.get(timeframe)
            fine_data = LazyFineData(symbol, fine_tf) if fine_tf else None

            strategies_data[fname] = {
                'symbol':     symbol,
                'timeframe':  timeframe,
                'data':       data,
                'fine_data':  fine_data,
                'smc_params': config.get('strategy', {}),
                'risk_params': config.get('risk', {}),
                'htf':        htf,
            }
        except Exception as e:
            print(f"  {Y}Fehler bei {fname}: {e}{NC}")
    return strategies_data


def _simulate_current_portfolio(settings: dict, strategies_data: dict,
                                 start_capital: float,
                                 start_date: str, end_date: str) -> dict | None:
    """Simuliert das aktuell aktive Portfolio auf dem gleichen Zeitraum."""
    from stbot.analysis.portfolio_simulator import run_portfolio_simulation
    current = [
        s for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
        if s.get('active')
    ]
    if not current:
        return None
    sim_data = {}
    for s in current:
        sym, tf = s.get('symbol', ''), s.get('timeframe', '')
        for fname, sd in strategies_data.items():
            if sd['symbol'] == sym and sd['timeframe'] == tf:
                sim_data[f"{sym}_{tf}"] = sd
                break
    if not sim_data:
        return None
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        return run_portfolio_simulation(start_capital, sim_data, start_date, end_date)


def _write_to_settings(portfolio_files: list, strategies_data: dict) -> None:
    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    existing     = settings.get('live_trading_settings', {}).get('active_strategies', [])
    existing_map = {(s.get('symbol'), s.get('timeframe')): s for s in existing}
    new_strategies = []
    for fname in portfolio_files:
        sd        = strategies_data.get(fname, {})
        symbol    = sd.get('symbol', '')
        timeframe = sd.get('timeframe', '')
        if not symbol or not timeframe:
            continue
        base  = existing_map.get((symbol, timeframe), {})
        entry = {**base, 'symbol': symbol, 'timeframe': timeframe, 'active': True}
        new_strategies.append(entry)
    lt = settings.setdefault('live_trading_settings', {})
    lt['active_strategies']          = new_strategies
    lt['use_auto_optimizer_results'] = False
    with open(SETTINGS_PATH, 'w') as f:
        json.dump(settings, f, indent=4)


BOT_NAME = 'stbot'


def _get_telegram_creds():
    try:
        with open(os.path.join(PROJECT_ROOT, 'secret.json')) as f:
            s = json.load(f)
        tg = s.get('telegram', {})
        t, c = tg.get('bot_token', ''), tg.get('chat_id', '')
        return (t, c) if t and c else (None, None)
    except Exception:
        return None, None


def _send_telegram(msg):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        requests.post(f'https://api.telegram.org/bot{token}/sendMessage',
                      data={'chat_id': chat, 'text': msg}, timeout=10)
    except Exception:
        pass


def _send_telegram_doc(fpath, caption=''):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        with open(fpath, 'rb') as fh:
            requests.post(f'https://api.telegram.org/bot{token}/sendDocument',
                          data={'chat_id': chat, 'caption': caption},
                          files={'document': fh}, timeout=30)
    except Exception:
        pass


def generate_trades_excel(final, strategies_data, capital, start_date, end_date):
    """Erstellt Excel-Tabelle mit allen Portfolio-Trades."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f'  {Y}openpyxl nicht installiert — Excel uebersprungen.{NC}')
        return None

    trades = final.get('trade_history', [])
    if not trades:
        return None

    fee_pct = 0.06 / 100  # muss zu portfolio_simulator.py passen (dort hart codiert)

    equity = capital
    rows = []
    for i, t in enumerate(trades, 1):
        pnl       = t.get('pnl', 0.0)
        entry     = t.get('entry', 0.0)
        exit_px   = t.get('exit', 0.0)
        direction = str(t.get('direction', 'long'))
        leverage  = t.get('leverage', 1) or 1
        margin    = t.get('margin_used', 0.0)
        notional  = margin * leverage
        move_pct  = ((exit_px / entry - 1) if direction == 'long' else (1 - exit_px / entry)) * 100 if entry else 0.0
        fee       = notional * fee_pct * 2
        equity += pnl
        rows.append({
            'Nr':                 i,
            'Datum':              str(t.get('entry_time', t.get('ts', '')))[:16].replace('T', ' '),
            'Coin':               str(t.get('symbol', '?')).split('/')[0],
            'Timeframe':          t.get('timeframe', '?'),
            'Richtung':           direction.upper(),
            'Ergebnis':           'TP erreicht' if pnl >= 0 else 'SL erreicht',
            'Reale Bewegung (%)': round(move_pct, 4),
            'Marge (USDT)':       round(margin, 4),
            'Gebühr (USDT)':      round(fee, 4),
            'PnL (USDT)':         round(pnl, 4),
            'Gesamtkapital':      round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'
    hdr  = PatternFill('solid', fgColor='1E3A5F')
    win  = PatternFill('solid', fgColor='D6F4DC')
    loss = PatternFill('solid', fgColor='FAD7D7')
    brd  = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                  top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    cw   = {'Nr': 6, 'Datum': 18, 'Coin': 10, 'Timeframe': 12, 'Richtung': 10,
             'Ergebnis': 14, 'Reale Bewegung (%)': 20, 'Marge (USDT)': 14,
             'Gebühr (USDT)': 14, 'PnL (USDT)': 14, 'Gesamtkapital': 16}
    hdrs = list(rows[0].keys())
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = brd
        ws.column_dimensions[get_column_letter(c)].width = cw.get(h, 14)
    ws.row_dimensions[1].height = 22
    for ri, row in enumerate(rows, 2):
        f = win if row['Ergebnis'] == 'TP erreicht' else loss
        for c, key in enumerate(hdrs, 1):
            cell = ws.cell(row=ri, column=c, value=row[key])
            cell.fill = f
            cell.border = brd
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Reale Bewegung (%)', 'Marge (USDT)', 'Gebühr (USDT)', 'PnL (USDT)', 'Gesamtkapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[ri].height = 18
    pnl = final.get('total_pnl_pct', 0)
    dd  = final.get('max_drawdown_pct', 0)
    wr  = final.get('win_rate', 0)
    eq  = final.get('end_capital', equity)
    n   = final.get('trade_count', len(trades))
    sr  = len(rows) + 3
    ws.cell(row=sr, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    sr += 1
    for label, val in [('Zeitraum', f'{start_date} -> {end_date}'), ('Trades', n),
                        ('Win-Rate', f'{wr:.1f}%'), ('PnL', f'{pnl:+.1f}%'),
                        ('Endkapital', f'{eq:.2f} USDT'), ('Max Drawdown', f'{dd:.1f}%')]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=val)
        sr += 1
    outfile = f'/tmp/{BOT_NAME}_trades.xlsx'
    wb.save(outfile)
    print(f'  {G}✓ Excel erstellt: {outfile}{NC}')
    return outfile


def generate_equity_html(final, capital, start_date, end_date, labels):
    """Erstellt interaktiven Portfolio-Equity-Chart (Optik wie dnabot:
    duenne Einzel-Symbol-Linien auf der primaeren Achse, dicke blaue
    Portfolio-Equity auf der sekundaeren Achse, Entry-/Exit-Marker)."""
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
        import pandas as pd
    except ImportError:
        print(f'  {Y}plotly/pandas nicht installiert — Chart uebersprungen.{NC}')
        return None

    eq_df = final.get('equity_curve')
    if eq_df is None or (hasattr(eq_df, 'empty') and eq_df.empty):
        return None

    trades   = final.get('trade_history', [])
    eq_times = pd.to_datetime(eq_df['timestamp'])
    eq_vals  = [float(v) for v in eq_df['equity']]
    pnl      = final.get('total_pnl_pct', 0)
    dd       = final.get('max_drawdown_pct', 0)
    wr       = final.get('win_rate', 0)
    n        = final.get('trade_count', 0)
    eq       = final.get('end_capital', eq_vals[-1] if eq_vals else capital)
    sign     = '+' if pnl >= 0 else ''
    title = (f"{BOT_NAME} Portfolio — {', '.join(labels)} | "
             f"PnL: {sign}{pnl:.1f}% | Equity: {eq:.2f} USDT | "
             f"MaxDD: {dd:.1f}% | WR: {wr:.1f}% | {n} Trades")

    # Nachschlage-Kurve fuer Marker-Positionen (naechster bekannter Portfolio-
    # Equity-Wert zu einem Trade-Zeitpunkt).
    eq_series = pd.Series(eq_vals, index=eq_times).sort_index()

    PAIR_COLORS = ['#f59e0b', '#8b5cf6', '#ec4899', '#14b8a6',
                   '#f97316', '#84cc16', '#06b6d4', '#a78bfa']

    fig = make_subplots(specs=[[{"secondary_y": True}]])

    # Einzel-Equity je Symbol/Timeframe (primaere Achse, duenn) -- eigene
    # Trade-Historie je Paar kumuliert, unabhaengig vom Gesamtportfolio.
    by_symbol = {}
    for t in trades:
        key = f"{t.get('symbol', '?')}/{t.get('timeframe', '?')}"
        by_symbol.setdefault(key, []).append(t)
    for idx, key in enumerate(sorted(by_symbol)):
        sym_trades = sorted(by_symbol[key], key=lambda t: str(t.get('entry_time', '')))
        peq    = capital
        ptimes = [sym_trades[0].get('entry_time', '')]
        pvals  = [peq]
        for t in sym_trades:
            peq += t.get('pnl', 0.0)
            ptimes.append(t.get('ts', t.get('entry_time', '')))
            pvals.append(round(peq, 2))
        fig.add_trace(go.Scatter(
            x=ptimes, y=pvals, mode='lines', name=key,
            line=dict(color=PAIR_COLORS[idx % len(PAIR_COLORS)], width=1),
            opacity=0.55,
        ), secondary_y=False)

    fig.add_hline(y=capital, line=dict(color='rgba(100,100,100,0.35)', width=1, dash='dash'),
                  annotation_text=f'Start {capital:.0f} USDT', annotation_position='top left')

    # Entry-/Exit-Marker auf der Portfolio-Equity (sekundaere Achse)
    entry_x, entry_y, entry_txt = [], [], []
    exit_win_x, exit_win_y   = [], []
    exit_loss_x, exit_loss_y = [], []
    for t in trades:
        try:
            y_entry = float(eq_series.asof(pd.Timestamp(t.get('entry_time'))))
            y_exit  = float(eq_series.asof(pd.Timestamp(t.get('ts'))))
        except Exception:
            continue
        entry_x.append(t.get('entry_time')); entry_y.append(y_entry)
        entry_txt.append(f"{t.get('symbol', '?')} {t.get('timeframe', '?')}<br>Equity: {y_entry:.2f} USDT")
        if t.get('pnl', 0.0) >= 0:
            exit_win_x.append(t.get('ts'));  exit_win_y.append(y_exit)
        else:
            exit_loss_x.append(t.get('ts')); exit_loss_y.append(y_exit)

    fig.add_trace(go.Scatter(x=list(eq_times), y=eq_vals, mode='lines', name='Portfolio Equity',
                             line=dict(color='#2563eb', width=2), opacity=0.75), secondary_y=True)

    if entry_x:
        fig.add_trace(go.Scatter(
            x=entry_x, y=entry_y, mode='markers',
            marker=dict(color='#16a34a', symbol='triangle-up', size=14, line=dict(width=1, color='#0f5132')),
            name='Entry ▲', text=entry_txt, hovertemplate='%{text}<extra>Entry</extra>',
        ), secondary_y=True)
    if exit_win_x:
        fig.add_trace(go.Scatter(
            x=exit_win_x, y=exit_win_y, mode='markers',
            marker=dict(color='#22d3ee', symbol='circle', size=11, line=dict(width=1, color='#0e7490')),
            name='Exit TP ✓',
        ), secondary_y=True)
    if exit_loss_x:
        fig.add_trace(go.Scatter(
            x=exit_loss_x, y=exit_loss_y, mode='markers',
            marker=dict(color='#ef4444', symbol='x', size=11, line=dict(width=2, color='#7f1d1d')),
            name='Exit SL ✗',
        ), secondary_y=True)

    fig.update_layout(
        title=dict(text=title, font=dict(size=13), x=0.5, xanchor='center'),
        height=750, hovermode='x unified', template='plotly_white', dragmode='zoom',
        xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
    )
    fig.update_yaxes(title_text='Einzel-Equity (USDT)', secondary_y=False, fixedrange=False)
    fig.update_yaxes(title_text='Portfolio-Equity (USDT)', secondary_y=True, fixedrange=False)

    outfile = f'/tmp/{BOT_NAME}_portfolio_equity.html'
    fig.write_html(outfile)
    print(f'  {G}✓ Chart erstellt: {outfile}{NC}')
    return outfile


def _do_replot(settings: dict, capital: float, start_date: str, end_date: str) -> int:
    print(f"\n{'─'*72}")
    print(f"{B}  stbot — Replot (aktives Portfolio){NC}")
    print(f"  Kapital: {capital:.0f} USDT | Zeitraum: {start_date} → {end_date}")
    print(f"{'─'*72}\n")

    active = [s for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
              if s.get('active')]
    if not active:
        print(f"{R}  Keine aktiven Strategien in settings.json.{NC}")
        return 1

    active_pairs = {(s['symbol'], s['timeframe']) for s in active}
    matching = []
    for path in _scan_configs():
        try:
            with open(path) as f:
                cfg = json.load(f)
            m = cfg.get('market', {})
            if (m.get('symbol'), m.get('timeframe')) in active_pairs:
                matching.append(path)
        except Exception:
            pass

    if not matching:
        print(f"{R}  Keine Config-Dateien fuer aktive Strategien gefunden.{NC}")
        return 1

    print(f"  {len(matching)} Config(s) gefunden.\n")
    strategies_data = _build_strategies_data(matching, start_date, end_date)
    if not strategies_data:
        print(f"{R}  Keine Daten geladen.{NC}")
        return 1

    from stbot.analysis.portfolio_simulator import run_portfolio_simulation
    sim_data = {f"{sd['symbol']}_{sd['timeframe']}": sd for sd in strategies_data.values()}
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        final = run_portfolio_simulation(capital, sim_data, start_date, end_date)
    if not final:
        print(f"{R}  Portfolio-Simulation fehlgeschlagen.{NC}")
        return 1

    selected_files = list(strategies_data.keys())
    labels = [f"{sd.get('symbol', '?')}/{sd.get('timeframe', '?')}"
              for sd in strategies_data.values()]
    pnl = final.get('total_pnl_pct', 0)
    dd  = final.get('max_drawdown_pct', 0)
    n   = final.get('trade_count', 0)
    wr  = final.get('win_rate', 0)
    eq  = final.get('end_capital', 0)

    print(f"\n{'='*72}")
    print(f"{B}  Replot — {len(selected_files)} Strategie(n){NC}\n")
    for fname, sd in strategies_data.items():
        print(f"  {G}✓{NC} {sd.get('symbol', fname):<26} / {sd.get('timeframe', ''):<6}")
    print(f"\n  Endkapital: {eq:.2f} USDT  | PnL: {pnl:+.1f}%  | MaxDD: {dd:.2f}%")
    print(f"{'='*72}\n")

    summary = (f"{BOT_NAME} Replot\n"
               f"{len(selected_files)} Strategien | {n} Trades | WR: {wr:.1f}%\n"
               f"PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}% | Equity: {eq:.2f} USDT\n"
               f"Zeitraum: {start_date} -> {end_date}")
    _send_telegram(summary)
    xlsx = generate_trades_excel(final, strategies_data, capital, start_date, end_date)
    if xlsx:
        _send_telegram_doc(xlsx, caption=f'{BOT_NAME} Trades | {n} Trades | WR: {wr:.1f}% | Equity: {eq:.2f} USDT')
    html = generate_equity_html(final, capital, start_date, end_date, labels)
    if html:
        _send_telegram_doc(html, caption=f'{BOT_NAME} Portfolio-Equity | PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}%')
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description='stbot Portfolio-Optimizer')
    parser.add_argument('--capital',    type=float, default=None)
    parser.add_argument('--max-dd',     type=float, default=30.0)
    parser.add_argument('--start-date', type=str,   default=None)
    parser.add_argument('--end-date',   type=str,   default=None)
    parser.add_argument('--auto-write', action='store_true')
    parser.add_argument('--replot',     action='store_true',
                        help='Replot fuer aktives Portfolio (keine Re-Optimierung)')
    args = parser.parse_args()

    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    opt           = settings.get('optimization_settings', {})
    capital       = args.capital or float(opt.get('start_capital', 100))
    max_dd        = args.max_dd
    end_date      = args.end_date   or date.today().strftime('%Y-%m-%d')
    # Startdatum-Fallback: backtest_lookback_weeks (rollend, hat Vorrang) ->
    # start_date (fixes Legacy-Datum) -> DEFAULT_LOOKBACK_DAYS.
    if args.start_date:
        start_date = args.start_date
    else:
        lookback_weeks = opt.get('backtest_lookback_weeks')
        if lookback_weeks:
            start_date = (date.today() - timedelta(weeks=int(lookback_weeks))).strftime('%Y-%m-%d')
        elif opt.get('start_date') and opt['start_date'] != 'auto':
            start_date = opt['start_date']
        else:
            start_date = (date.today() - timedelta(days=DEFAULT_LOOKBACK_DAYS)).strftime('%Y-%m-%d')
    max_positions = int(settings.get('live_trading_settings', {}).get('max_open_positions', 10))

    if args.replot:
        return _do_replot(settings, capital, start_date, end_date)

    print(f"\n{'─'*72}")
    print(f"{B}  stbot — Automatische Portfolio-Optimierung{NC}")
    print(f"  Greedy-Selektion mit echter Portfolio-Simulation (MaxDD ≤ {max_dd:.0f}%)")
    print(f"  Kapital: {capital:.0f} USDT | Positionen: max {max_positions} | "
          f"Zeitraum: {start_date} → {end_date}")
    print(f"{'─'*72}\n")

    config_files = _scan_configs()
    if not config_files:
        print(f"{R}  Keine Configs in {CONFIGS_DIR}{NC}")
        print(f"  → Zuerst run_pipeline.sh ausfuehren!\n")
        return 1

    print(f"  {len(config_files)} Config(s) gefunden.\n")
    strategies_data = _build_strategies_data(config_files, start_date, end_date)
    if not strategies_data:
        print(f"{R}  Keine Daten geladen.{NC}")
        return 1

    from stbot.analysis.portfolio_optimizer import run_portfolio_optimizer
    result = run_portfolio_optimizer(capital, strategies_data, start_date, end_date, max_dd)

    if not result or not result.get('optimal_portfolio'):
        print(f"{R}  Kein Portfolio erfuellt die Bedingungen (MaxDD ≤ {max_dd:.0f}%).{NC}\n")
        return 0

    portfolio_files = result['optimal_portfolio'][:max_positions]
    final           = result.get('final_result') or {}

    print(f"\n{'='*72}")
    print(f"{B}  Optimales Portfolio — {len(portfolio_files)} Strategie(n){NC}\n")
    for fname in portfolio_files:
        sd = strategies_data.get(fname, {})
        print(f"  {G}✓{NC} {sd.get('symbol', fname):<26} / {sd.get('timeframe', ''):<6}")
    if final:
        pnl = final.get('total_pnl_pct', 0)
        print(f"\n  Endkapital: {final.get('end_capital', 0):.2f} USDT  "
              f"| PnL: {pnl:+.1f}%  "
              f"| MaxDD: {final.get('max_drawdown_pct', 0):.2f}%")
    print(f"{'='*72}\n")

    current_set = {
        (s.get('symbol'), s.get('timeframe'))
        for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
        if s.get('active')
    }
    new_set = {
        (strategies_data.get(f, {}).get('symbol'), strategies_data.get(f, {}).get('timeframe'))
        for f in portfolio_files
    }

    cur_result  = _simulate_current_portfolio(settings, strategies_data, capital, start_date, end_date)
    cur_cap     = cur_result.get('end_capital', 0) if cur_result else 0
    new_cap     = final.get('end_capital', 0)
    if cur_result:
        print(f"  Aktuelles Portfolio: {cur_cap:.2f} USDT  "
              f"| PnL: {cur_result.get('total_pnl_pct', 0):+.1f}%  "
              f"| MaxDD: {cur_result.get('max_drawdown_pct', 0):.2f}%")
        print(f"  Neues Portfolio:     {new_cap:.2f} USDT  "
              f"| PnL: {final.get('total_pnl_pct', 0):+.1f}%  "
              f"| MaxDD: {final.get('max_drawdown_pct', 0):.2f}%\n")

    if args.auto_write:
        if cur_result and new_cap <= cur_cap:
            print(f"{Y}  Neues Portfolio ({new_cap:.2f} USDT) nicht besser als aktuelles "
                  f"({cur_cap:.2f} USDT) — keine Aenderung.{NC}\n")
        else:
            _write_to_settings(portfolio_files, strategies_data)
            print(f"{G}✓ settings.json aktualisiert — {len(portfolio_files)} Strategie(n).{NC}\n")
    else:
        if current_set == new_set:
            print(f"{Y}  Portfolio unveraendert — keine Aenderung noetig.{NC}\n")
        else:
            try:
                ans = input("  Optimales Portfolio in settings.json eintragen? (j/n): ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                ans = 'n'
            if ans in ('j', 'ja', 'y', 'yes'):
                _write_to_settings(portfolio_files, strategies_data)
                print(f"{G}✓ settings.json aktualisiert.{NC}\n")
            else:
                print(f"{Y}  settings.json NICHT geaendert.{NC}\n")

    # ── Reports & Telegram ──────────────────────────────────────────────────
    if args.auto_write:
        labels = [
            f"{strategies_data.get(f, {}).get('symbol', '?')}/{strategies_data.get(f, {}).get('timeframe', '?')}"
            for f in portfolio_files
        ]
        pnl = final.get('total_pnl_pct', 0)
        dd  = final.get('max_drawdown_pct', 0)
        n   = final.get('trade_count', 0)
        wr  = final.get('win_rate', 0)
        eq  = final.get('end_capital', 0)
        summary = (f"{BOT_NAME} Auto-Optimizer\n"
                   f"{len(portfolio_files)} Strategien | {n} Trades | WR: {wr:.1f}%\n"
                   f"PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}% | Equity: {eq:.2f} USDT\n"
                   f"Zeitraum: {start_date} -> {end_date}")
        _send_telegram(summary)
        xlsx = generate_trades_excel(final, strategies_data, capital, start_date, end_date)
        if xlsx:
            _send_telegram_doc(xlsx, caption=f'{BOT_NAME} Trades | {n} Trades | WR: {wr:.1f}% | Equity: {eq:.2f} USDT')
        html = generate_equity_html(final, capital, start_date, end_date, labels)
        if html:
            _send_telegram_doc(html, caption=f'{BOT_NAME} Portfolio-Equity | PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}%')

    return 0


if __name__ == '__main__':
    sys.exit(main())
