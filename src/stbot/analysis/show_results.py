# /root/stbot/src/stbot/analysis/show_results.py
import os
import sys
import json
import pandas as pd
from datetime import date
import logging
import argparse

logging.getLogger('tensorflow').setLevel(logging.ERROR)
logging.getLogger('absl').setLevel(logging.ERROR)
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='keras')

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.append(os.path.join(PROJECT_ROOT, 'src'))

# KORREKTUR: Import run_backtest statt run_smc_backtest
from stbot.analysis.backtester import load_data, run_backtest
from stbot.analysis.portfolio_simulator import run_portfolio_simulation
from stbot.analysis.portfolio_optimizer import run_portfolio_optimizer
from stbot.utils.telegram import send_document

# --- Einzel-Analyse ---
def run_single_analysis(start_date, end_date, start_capital):
    print("--- StBot Ergebnis-Analyse (Einzel-Modus) ---")
    configs_dir = os.path.join(PROJECT_ROOT, 'src', 'stbot', 'strategy', 'configs')
    all_results = []
    
    if not os.path.exists(configs_dir):
        print(f"Konfigurationsverzeichnis nicht gefunden: {configs_dir}")
        return

    config_files = sorted([f for f in os.listdir(configs_dir) if f.startswith('config_') and f.endswith('.json')])
    
    if not config_files:
        print("\nKeine gültigen Konfigurationen zum Analysieren gefunden."); return
    
    print(f"Zeitraum: {start_date} bis {end_date} | Startkapital: {start_capital} USDT")
    
    for filename in config_files:
        config_path = os.path.join(configs_dir, filename)
        if not os.path.exists(config_path): continue
        try:
            with open(config_path, 'r') as f: config = json.load(f)
            symbol = config['market']['symbol']
            timeframe = config['market']['timeframe']
            strategy_name = f"{symbol} ({timeframe})"
            
            print(f"\nAnalysiere Ergebnisse für: {filename}...")
            data = load_data(symbol, timeframe, start_date, end_date)
            if data.empty:
                print(f"--> WARNUNG: Konnte keine Daten laden für {strategy_name}. Überspringe."); continue

            strategy_params = config.get('strategy', {})
            risk_params = config.get('risk', {})

            # Parameter für den Backtester vorbereiten
            strategy_params['symbol'] = symbol
            strategy_params['timeframe'] = timeframe
            strategy_params['htf'] = config['market'].get('htf')

            # KORREKTUR: Aufruf von run_backtest statt run_smc_backtest
            result = run_backtest(data.copy(), strategy_params, risk_params, start_capital, verbose=False)
            
            all_results.append({
                "Strategie": strategy_name,
                "Trades": result.get('trades_count', 0),
                "Win Rate %": result.get('win_rate', 0),
                "PnL %": result.get('total_pnl_pct', -100),
                "Max DD %": result.get('max_drawdown_pct', 1.0) * 100,
                "Endkapital": result.get('end_capital', start_capital)
            })
        except Exception as e:
            print(f"--> FEHLER bei der Analyse von {filename}: {e}")
            continue
            
    if not all_results:
        print("\nKeine gültigen Ergebnisse zum Anzeigen gefunden."); return
        
    results_df = pd.DataFrame(all_results)
    results_df = results_df.sort_values(by="PnL %", ascending=False)
    
    pd.set_option('display.width', 1000); pd.set_option('display.max_columns', None)
    print("\n\n=========================================================================================");
    print(f"                        Zusammenfassung aller Einzelstrategien");
    print("=========================================================================================")
    pd.set_option('display.float_format', '{:.2f}'.format);
    print(results_df.to_string(index=False));
    print("=========================================================================================")


# --- Geteilter Modus (Manuell / Auto) ---
def run_shared_mode(is_auto: bool, start_date, end_date, start_capital, target_max_dd: float):
    mode_name = "Automatische Portfolio-Optimierung" if is_auto else "Manuelle Portfolio-Simulation"
    print(f"--- StBot {mode_name} ---")
    if is_auto:
        print(f"Ziel: Maximaler Profit bei maximal {target_max_dd:.2f}% Drawdown.")

    configs_dir = os.path.join(PROJECT_ROOT, 'src', 'stbot', 'strategy', 'configs')
    available_strategies = []
    if os.path.isdir(configs_dir):
        for filename in sorted(os.listdir(configs_dir)):
            if filename.startswith('config_') and filename.endswith('.json'):
                available_strategies.append(filename)
    
    if not available_strategies:
        print("Keine optimierten Strategien (Configs) gefunden."); return

    selected_files = []
    if not is_auto:
        print("\nVerfügbare Strategien:")
        for i, name in enumerate(available_strategies): print(f"  {i+1}) {name}")
        selection = input("\nWelche Strategien sollen simuliert werden? (Zahlen mit Komma, z.B. 1,3,4 oder 'alle'): ")
        try:
            if selection.lower() == 'alle': selected_files = available_strategies
            else: selected_files = [available_strategies[int(i.strip()) - 1] for i in selection.split(',')]
        except (ValueError, IndexError): print("Ungültige Auswahl. Breche ab."); return
    else:
        selected_files = available_strategies

    strategies_data = {}
    print("\nLade Daten für gewählte Strategien...")
    for filename in selected_files:
        try:
            with open(os.path.join(configs_dir, filename), 'r') as f: config = json.load(f)
            symbol = config['market']['symbol']
            timeframe = config['market']['timeframe']
            htf = config['market'].get('htf')

            data = load_data(symbol, timeframe, start_date, end_date)
            if not data.empty:
                strategies_data[filename] = {
                    'symbol': symbol, 'timeframe': timeframe, 'data': data,
                    'smc_params': config.get('strategy', {}), # Name smc_params wird im Simulator noch erwartet
                    'risk_params': config.get('risk', {}),
                    'htf': htf
                }
            else:
                print(f"WARNUNG: Konnte Daten für {filename} nicht laden. Wird ignoriert.")
        except Exception as e:
            print(f"FEHLER beim Laden der Config/Daten für {filename}: {e}")

    if not strategies_data:
        print("Konnte für keine der gewählten Strategien Daten laden. Breche ab."); return

    equity_df = pd.DataFrame()
    csv_path = ""
    caption = ""

    try:
        if is_auto:
            results = run_portfolio_optimizer(start_capital, strategies_data, start_date, end_date, target_max_dd)

            if results and 'final_result' in results and results['final_result'] is not None:
                final_report = results['final_result']
                print("\n======================================================="); print("     Ergebnis der automatischen Portfolio-Optimierung"); print("=======================================================")
                print(f"Zeitraum: {start_date} bis {end_date}\nStartkapital: {start_capital:.2f} USDT")
                print(f"Bedingung: Max Drawdown <= {target_max_dd:.2f}%")

                if results.get('optimal_portfolio'):
                    print("\nOptimales Portfolio gefunden (" + str(len(results['optimal_portfolio'])) + " Strategien):")
                    for strat_filename in results['optimal_portfolio']: print(f"  - {strat_filename}")
                else:
                    print("\nBeste Einzelstrategie gefunden:")
                    strat_key = final_report.get('strategy_key', 'Unbekannt')
                    print(f"  - {strat_key}")

                print("\n--- Simulierte Performance dieses Portfolios/dieser Strategie ---")
                print(f"Endkapital:         {final_report['end_capital']:.2f} USDT"); print(f"Gesamt PnL:         {final_report['end_capital'] - start_capital:+.2f} USDT ({final_report['total_pnl_pct']:.2f}%)")
                print(f"Portfolio Max DD:   {final_report['max_drawdown_pct']:.2f}%")
                liq_date = final_report.get('liquidation_date')
                print(f"Liquidiert:         {'JA, am ' + liq_date.strftime('%Y-%m-%d') if liq_date else 'NEIN'}")

                csv_path = os.path.join(PROJECT_ROOT, 'optimal_portfolio_equity.csv')
                caption = f"Automatischer Portfolio-Optimierungsbericht (Max DD <= {target_max_dd:.1f}%)\nEndkapital: {final_report['end_capital']:.2f} USDT"
                equity_df = final_report.get('equity_curve')
            else:
                print(f"\nKein Portfolio gefunden, das die Bedingung Max Drawdown <= {target_max_dd:.2f}% erfüllt.")

        # --- Manuelle Simulation ---
        else:
            sim_data = {v['symbol'] + "_" + v['timeframe']: v for k, v in strategies_data.items()}
            results = run_portfolio_simulation(start_capital, sim_data, start_date, end_date)
            if results:
                print("\n======================================================="); print("           Portfolio-Simulations-Ergebnis"); print("=======================================================")
                print(f"Zeitraum: {start_date} bis {end_date}\nStartkapital: {results['start_capital']:.2f} USDT")
                print("\n--- Gesamt-Performance ---")
                print(f"Endkapital:         {results['end_capital']:.2f} USDT"); print(f"Gesamt PnL:         {results['end_capital'] - results['start_capital']:+.2f} USDT ({results['total_pnl_pct']:.2f}%)")
                print(f"Anzahl Trades:       {results['trade_count']}"); print(f"Win-Rate:           {results['win_rate']:.2f}%")
                print(f"Portfolio Max DD:   {results['max_drawdown_pct']:.2f}% am {results['max_drawdown_date'].strftime('%Y-%m-%d') if results['max_drawdown_date'] else 'N/A'}")
                liq_date = results.get('liquidation_date')
                print(f"Liquidiert:         {'JA, am ' + liq_date.strftime('%Y-%m-%d') if liq_date else 'NEIN'}")

                csv_path = os.path.join(PROJECT_ROOT, 'manual_portfolio_equity.csv')
                caption = f"Manueller Portfolio-Simulationsbericht\nEndkapital: {results['end_capital']:.2f} USDT"
                equity_df = results.get('equity_curve')

    except Exception as e:
        print(f"\nFEHLER während der Portfolio-Analyse: {e}")
        import traceback
        traceback.print_exc()
        equity_df = pd.DataFrame()

    # --- Ergebnisse speichern und senden ---
    if equity_df is not None and not equity_df.empty and csv_path:
        print("\n--- Export ---")
        try:
            export_cols = ['timestamp', 'equity', 'drawdown_pct']
            available_cols = [col for col in export_cols if col in equity_df.columns]
            if 'timestamp' in equity_df.columns and not isinstance(equity_df.index, pd.DatetimeIndex):
                equity_df['timestamp'] = pd.to_datetime(equity_df['timestamp'])
                equity_df.set_index('timestamp', inplace=True, drop=False)

            equity_df_export = equity_df[available_cols].copy()
            if 'timestamp' not in equity_df_export.columns and isinstance(equity_df.index, pd.DatetimeIndex):
                equity_df_export.insert(0, 'timestamp', equity_df.index)

            equity_df_export.to_csv(csv_path, index=False)
            print(f"✔ Details zur Equity-Kurve wurden nach '{os.path.basename(csv_path)}' exportiert.")
        except Exception as e_csv:
            print(f"FEHLER beim Speichern der CSV '{csv_path}': {e_csv}")

        # Chart & Excel
        final_sim_for_export = None
        portfolio_files_for_export = []
        if is_auto and 'results' in dir() and results and 'final_result' in results:
            final_sim_for_export    = results['final_result']
            portfolio_files_for_export = results.get('optimal_portfolio', [])
        elif not is_auto and 'results' in dir() and results:
            final_sim_for_export    = results
            portfolio_files_for_export = selected_files

        if final_sim_for_export:
            do_export = is_auto
            if not is_auto:
                ans = input("\n  Charts & Excel erstellen und via Telegram senden? (j/n) [Standard: n]: ").strip().lower()
                do_export = ans in ('j', 'y', 'ja')
            if do_export:
                _generate_portfolio_chart(final_sim_for_export, portfolio_files_for_export,
                                          start_capital, start_date, end_date)
                _generate_trades_excel(final_sim_for_export, start_capital)

        print("=======================================================")

    elif csv_path:
        print(f"\nKeine Equity-Daten zum Exportieren für '{os.path.basename(csv_path)}' vorhanden.")
    else:
        print("\nPortfolio-Analyse fehlgeschlagen oder kein gültiges Portfolio gefunden, kein Export möglich.")


GREEN  = '\033[0;32m'
YELLOW = '\033[1;33m'
NC     = '\033[0m'


def _get_telegram_cfg():
    try:
        with open(os.path.join(PROJECT_ROOT, 'secret.json'), 'r') as f:
            s = json.load(f)
        tg = s.get('telegram', {})
        return tg.get('bot_token', ''), tg.get('chat_id', '')
    except Exception:
        return '', ''


def _generate_portfolio_chart(final_sim, portfolio_files, capital, start_date, end_date):
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print(f"  {YELLOW}plotly nicht installiert — Chart übersprungen. (pip install plotly){NC}")
        return

    eq_df         = final_sim.get('equity_curve')
    trade_history = final_sim.get('trade_history', [])
    if eq_df is None or (hasattr(eq_df, 'empty') and eq_df.empty):
        print(f"  {YELLOW}Keine Equity-Daten — Chart übersprungen.{NC}")
        return

    eq_times = eq_df['timestamp'].astype(str).tolist()
    eq_vals  = eq_df['equity'].tolist()

    win_x, win_y   = [], []
    loss_x, loss_y = [], []
    for t in trade_history:
        ts  = str(t.get('ts', ''))
        row = eq_df[eq_df['timestamp'] <= pd.to_datetime(t['ts'])]
        eq_at = float(row['equity'].iloc[-1]) if not row.empty else capital
        if float(t['pnl']) > 0:
            win_x.append(ts);  win_y.append(eq_at)
        else:
            loss_x.append(ts); loss_y.append(eq_at)

    n_strats  = len(portfolio_files)
    pairs = []
    for fname in portfolio_files:
        name  = fname.replace('config_', '').replace('.json', '')
        parts = name.split('_')
        tf    = parts[-1] if parts else ''
        sym   = parts[0][:3].upper() if parts else ''
        pairs.append(f"{sym}/{tf}")
    pairs_str = ', '.join(pairs)

    pnl_pct = final_sim.get('total_pnl_pct', 0)
    sign    = '+' if pnl_pct >= 0 else ''
    title = (
        f"StBot Portfolio — {n_strats} Strategie(n) ({pairs_str}) | "
        f"Zeitraum: {start_date} → {end_date} | "
        f"Trades: {final_sim.get('trade_count', 0)} | WR: {final_sim.get('win_rate', 0):.1f}% | "
        f"PnL: {sign}{pnl_pct:.1f}% | "
        f"Endkapital: {final_sim.get('end_capital', capital):.2f} USDT | "
        f"MaxDD: {final_sim.get('max_drawdown_pct', 0):.1f}%"
    )

    fig = make_subplots(specs=[[{"secondary_y": False}]])
    fig.add_hline(
        y=capital,
        line=dict(color='rgba(100,100,100,0.35)', width=1, dash='dash'),
        annotation_text=f'Start {capital:.0f} USDT',
        annotation_position='top left',
    )

    STRAT_COLORS = [
        '#f59e0b', '#10b981', '#8b5cf6', '#f97316',
        '#ec4899', '#14b8a6', '#a3e635', '#fb923c',
        '#e879f9', '#38bdf8',
    ]
    strat_trades = {}
    for t in trade_history:
        sym = t.get('symbol', '').split('/')[0]
        tf  = t.get('timeframe', '')
        k   = f"{sym}/{tf}"
        strat_trades.setdefault(k, []).append(t)

    for idx, (strat_key, trades) in enumerate(sorted(strat_trades.items())):
        trades_sorted = sorted(trades, key=lambda x: x.get('ts', ''))
        eq = capital
        xs = [str(trades_sorted[0].get('ts', ''))[:16]]
        ys = [capital]
        for t in trades_sorted:
            eq += float(t['pnl'])
            xs.append(str(t.get('ts', ''))[:16])
            ys.append(round(eq, 4))
        color = STRAT_COLORS[idx % len(STRAT_COLORS)]
        fig.add_trace(go.Scatter(
            x=xs, y=ys, mode='lines', name=strat_key,
            line=dict(color=color, width=1.2, dash='dot'),
            opacity=0.6,
            hovertemplate=f"{strat_key}: %{{y:.2f}} USDT<extra></extra>",
        ))

    fig.add_trace(go.Scatter(
        x=eq_times, y=eq_vals, mode='lines', name='Portfolio Equity',
        line=dict(color='#2563eb', width=2.5),
        hovertemplate='Portfolio: %{y:.2f} USDT<extra></extra>',
    ))

    if win_x:
        fig.add_trace(go.Scatter(
            x=win_x, y=win_y, mode='markers',
            marker=dict(color='#22d3ee', symbol='circle', size=8,
                        line=dict(width=1, color='#0e7490')),
            name='TP ✓',
        ))
    if loss_x:
        fig.add_trace(go.Scatter(
            x=loss_x, y=loss_y, mode='markers',
            marker=dict(color='#ef4444', symbol='x', size=8,
                        line=dict(width=2, color='#7f1d1d')),
            name='SL ✗',
        ))

    fig.update_layout(
        title=dict(text=title, font=dict(size=12), x=0.5, xanchor='center'),
        height=600, hovermode='x unified', template='plotly_dark', dragmode='zoom',
        xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
        margin=dict(l=60, r=60, t=80, b=40),
        yaxis=dict(title='Equity (USDT)', fixedrange=False),
    )

    out_dir  = os.path.join(PROJECT_ROOT, 'artifacts', 'charts')
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, 'stbot_portfolio_equity.html')
    fig.write_html(out_file)
    print(f"  {GREEN}Chart gespeichert: stbot_portfolio_equity.html{NC}")

    bot_token, chat_id = _get_telegram_cfg()
    if bot_token and chat_id:
        caption = (
            f"StBot Portfolio-Equity\n"
            f"{start_date} → {end_date} | {n_strats} Strategie(n) | "
            f"PnL: {sign}{pnl_pct:.1f}% | Equity: {final_sim.get('end_capital', capital):.2f} USDT | "
            f"MaxDD: {final_sim.get('max_drawdown_pct', 0):.1f}%"
        )
        send_document(bot_token, chat_id, out_file, caption=caption)
        print(f"  {GREEN}Chart via Telegram gesendet.{NC}")
    else:
        print(f"  {YELLOW}Telegram nicht konfiguriert — nur lokal gespeichert.{NC}")


def _generate_trades_excel(final_sim, capital):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  {YELLOW}openpyxl nicht installiert — Excel übersprungen. (pip install openpyxl){NC}")
        return

    trade_history = final_sim.get('trade_history', [])
    if not trade_history:
        print(f"  {YELLOW}Keine Trades — Excel übersprungen.{NC}")
        return

    equity = capital
    rows   = []
    for i, t in enumerate(trade_history):
        pnl      = float(t['pnl'])
        equity  += pnl
        sym      = t.get('symbol', '')
        tf       = t.get('timeframe', '')
        strat    = f"{sym.split('/')[0]}/{tf}" if sym else tf
        dir_     = t.get('direction', '').upper()
        entry    = round(float(t.get('entry', 0)), 6)
        exit_p   = round(float(t.get('exit',  0)), 6)
        ergebnis = 'TP erreicht' if pnl > 0 else 'SL erreicht'
        lev      = float(t.get('leverage', 1) or 1)
        margin   = float(t.get('margin_used', 0))
        if entry > 0:
            raw_move = (exit_p - entry) / entry * 100.0
            move_pct = raw_move if dir_ == 'LONG' else -raw_move
        else:
            move_pct = 0.0
        rows.append({
            'Nr':                 i + 1,
            'Datum':              str(t.get('entry_time', t.get('ts', '')))[:16].replace('T', ' '),
            'Strategie':          strat,
            'Richtung':           dir_,
            'Hebel':              int(lev) if lev else '—',
            'Reale Bewegung (%)': round(move_pct, 4),
            'Marge (USDT)':       round(margin, 4),
            'Entry':              entry,
            'Exit':               exit_p,
            'Ergebnis':           ergebnis,
            'PnL (USDT)':         round(pnl,    4),
            'Kapital':            round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    win_fill    = PatternFill('solid', fgColor='D6F4DC')
    loss_fill   = PatternFill('solid', fgColor='FAD7D7')
    alt_fill    = PatternFill('solid', fgColor='F2F2F2')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    col_widths = {
        'Nr': 5, 'Datum': 18, 'Strategie': 22, 'Richtung': 10,
        'Hebel': 8, 'Reale Bewegung (%)': 18, 'Marge (USDT)': 14,
        'Entry': 14, 'Exit': 14, 'Ergebnis': 14, 'PnL (USDT)': 14, 'Kapital': 16,
    }

    headers = list(rows[0].keys())
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        if row['Ergebnis'] == 'TP erreicht':
            fill = win_fill
        elif r_idx % 2 == 0:
            fill = loss_fill
        else:
            fill = alt_fill
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Entry', 'Exit', 'PnL (USDT)', 'Kapital', 'Reale Bewegung (%)', 'Marge (USDT)'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[r_idx].height = 18

    total = len(rows)
    wins  = sum(1 for r in rows if r['Ergebnis'] == 'TP erreicht')
    sr    = total + 3
    pnl_total = rows[-1]['Kapital'] - capital if rows else 0.0
    pnl_pct   = pnl_total / capital * 100 if capital else 0.0
    ws.cell(row=sr, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    for label, value in [
        ('Trades gesamt', total),
        ('Win-Rate',      f"{wins / total * 100:.1f}%" if total else '—'),
        ('PnL',           f"{pnl_pct:+.1f}%"),
        ('Endkapital',    f"{rows[-1]['Kapital']:.2f} USDT" if rows else '—'),
    ]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=value)
        sr += 1

    out_dir  = os.path.join(PROJECT_ROOT, 'artifacts', 'charts')
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, 'stbot_trades.xlsx')
    wb.save(out_file)
    print(f"  {GREEN}Excel gespeichert: stbot_trades.xlsx{NC}")

    bot_token, chat_id = _get_telegram_cfg()
    if bot_token and chat_id:
        caption = (f"StBot Trades — {total} Trades | "
                   f"WR: {wins / total * 100:.1f}% | PnL: {pnl_pct:+.1f}%" if total else "StBot Trades")
        send_document(bot_token, chat_id, out_file, caption=caption)
        print(f"  {GREEN}Excel via Telegram gesendet.{NC}")
    else:
        print(f"  {YELLOW}Telegram nicht konfiguriert — nur lokal gespeichert.{NC}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--mode', default='1', type=str,
                        choices=['1', '2', '3', '4'],
                        help="Analyse-Modus: 1=Einzel, 2=Manuell, 3=Auto, 4=Interaktive Charts")
    parser.add_argument('--target_max_drawdown', default=30.0, type=float,
                        help="Ziel Max Drawdown %% (nur fuer Modus 3)")
    args = parser.parse_args()

    # Mode 4 hat eigenes Input-System
    if args.mode == '4':
        try:
            from stbot.analysis.interactive_status import main as interactive_main
            interactive_main()
        except Exception as e:
            print(f"Fehler beim Ausfuehren der interaktiven Charts: {e}")
            import traceback
            traceback.print_exc()
        sys.exit(0)

    print("\n--- Bitte Konfiguration für den Backtest festlegen ---")
    start_date = input(f"Startdatum (JJJJ-MM-TT) [Standard: 2023-01-01]: ") or "2023-01-01"
    end_date = input(f"Enddatum (JJJJ-MM-TT) [Standard: Heute]: ") or date.today().strftime("%Y-%m-%d")
    start_capital = int(input(f"Startkapital in USDT eingeben [Standard: 1000]: ") or 1000)
    print("--------------------------------------------------")

    if args.mode == '2':
        run_shared_mode(
            is_auto=False,
            start_date=start_date,
            end_date=end_date,
            start_capital=start_capital,
            target_max_dd=999.0
        )
    elif args.mode == '3':
        run_shared_mode(
            is_auto=True,
            start_date=start_date,
            end_date=end_date,
            start_capital=start_capital,
            target_max_dd=args.target_max_drawdown
        )
    else:
        run_single_analysis(start_date=start_date, end_date=end_date, start_capital=start_capital)
